#!/usr/bin/env python3
"""
🤖 SHP 파일 배치 AI 분석기
독립적으로 실행되는 PyQt GUI 애플리케이션

기능:
- 폴더 내 모든 SHP 파일 배치 분석
- 좌표계 자동 감지 및 수동 선택
- AI 점 개수 vs 실제 점 개수 비교
- 실시간 진행 모니터링
- 결과 엑셀 저장
"""

import sys
import os
import gc
import time
import logging
import subprocess
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import geopandas as gpd
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QGroupBox, QLabel, QPushButton, QComboBox, QProgressBar, QTableWidget,
    QTableWidgetItem, QHeaderView, QDialog, QRadioButton, QCheckBox,
    QFileDialog, QMessageBox
)
from PyQt5.QtCore import QThread, pyqtSignal, Qt
try:
    from PyQt5.QtCore import QCoreApplication
except ImportError:
    QCoreApplication = None
from PyQt5.QtGui import QFont, QColor

# 프로젝트 경로 추가
current_dir = Path(__file__).parent
sys.path.insert(0, str(current_dir))
sys.path.insert(0, str(current_dir / "src"))

# 가능한 경로들 시도
possible_paths = [
    current_dir,
    current_dir / "src",
    current_dir.parent,
    current_dir.parent / "src",
]

for path in possible_paths:
    sys.path.insert(0, str(path))

# 기존 모듈 import
PipelineManager = None
DistrictRoadClipper = None
SkeletonExtractor = None
has_full_modules = False

# 여러 경로에서 모듈 찾기 시도
import_attempts = [
    ("src.core.pipeline_manager", "PipelineManager"),
    ("src.core.district_road_clipper", "DistrictRoadClipper"),
    ("src.core.skeleton_extractor", "SkeletonExtractor"),
]

# 첫 번째 시도: src.core 경로
try:
    import importlib
    PipelineManager = getattr(importlib.import_module("src.core.pipeline_manager"), "PipelineManager")
    DistrictRoadClipper = getattr(importlib.import_module("src.core.district_road_clipper"), "DistrictRoadClipper")
    SkeletonExtractor = getattr(importlib.import_module("src.core.skeleton_extractor"), "SkeletonExtractor")
    has_full_modules = True
    print("✅ 프로젝트 모듈 로드 성공 (src.core)")
except (ImportError, AttributeError, ModuleNotFoundError):
    # 두 번째 시도: core 직접 경로
    try:
        PipelineManager = getattr(importlib.import_module("core.pipeline_manager"), "PipelineManager")
        DistrictRoadClipper = getattr(importlib.import_module("core.district_road_clipper"), "DistrictRoadClipper")
        SkeletonExtractor = getattr(importlib.import_module("core.skeleton_extractor"), "SkeletonExtractor")
        has_full_modules = True
        print("✅ 코어 모듈 로드 성공 (core)")
    except (ImportError, AttributeError, ModuleNotFoundError):
        print("⚠️ 기존 프로젝트 모듈을 찾을 수 없습니다.")
        print("대신 기본 분석 기능을 사용합니다.")
        print("\n📁 권장 파일 구조:")
        print("- models/ 폴더 (AI 모델 파일)")
        print("- src/core/ 폴더 (전체 기능)")
        print("\n🔧 현재는 기본 기능으로 실행됩니다.")

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class CRSSelectionDialog(QDialog):
    """좌표계 선택 다이얼로그"""
    
    def __init__(self, shp_file: str, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.shp_file = shp_file
        self.selected_crs: Optional[str] = None
        self.apply_to_all = False
        self.skip_file = False
        self.init_ui()
    
    def init_ui(self) -> None:
        """UI 초기화"""
        self.setWindowTitle("좌표계 선택 필요")
        self.setModal(True)
        self.setFixedSize(500, 300)
        
        layout = QVBoxLayout()
        
        # 파일 정보
        info_label = QLabel(f"⚠️ 좌표계 자동 감지 실패\n파일: {Path(self.shp_file).name}")
        info_label.setStyleSheet("QLabel {background-color: #fff3cd; padding: 10px; border-radius: 5px;}")
        layout.addWidget(info_label)
        
        # 좌표계 선택
        crs_group = QGroupBox("좌표계 선택")
        crs_layout = QVBoxLayout()
        
        self.crs_5186_radio = QRadioButton("EPSG:5186 (중부원점) - 서울, 경기, 충청, 전라, 제주")
        self.crs_5186_radio.setChecked(True)
        crs_layout.addWidget(self.crs_5186_radio)
        
        self.crs_5187_radio = QRadioButton("EPSG:5187 (동부원점) - 부산, 대구, 울산, 경상")
        crs_layout.addWidget(self.crs_5187_radio)
        
        crs_group.setLayout(crs_layout)
        layout.addWidget(crs_group)
        
        # 옵션
        option_layout = QHBoxLayout()
        self.apply_all_checkbox = QCheckBox("남은 파일에도 동일하게 적용")
        option_layout.addWidget(self.apply_all_checkbox)
        layout.addLayout(option_layout)
        
        # 버튼
        button_layout = QHBoxLayout()
        
        apply_btn = QPushButton("적용")
        apply_btn.setStyleSheet("QPushButton {background-color: #28a745; color: white; padding: 8px;}")
        apply_btn.clicked.connect(self.accept_crs)
        button_layout.addWidget(apply_btn)
        
        skip_btn = QPushButton("건너뛰기")
        skip_btn.setStyleSheet("QPushButton {background-color: #ffc107; color: black; padding: 8px;}")
        skip_btn.clicked.connect(self.skip_current)
        button_layout.addWidget(skip_btn)
        
        cancel_btn = QPushButton("취소")
        cancel_btn.setStyleSheet("QPushButton {background-color: #dc3545; color: white; padding: 8px;}")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
    
    def accept_crs(self) -> None:
        """좌표계 적용"""
        if self.crs_5187_radio.isChecked():
            self.selected_crs = 'EPSG:5187'
        else:
            self.selected_crs = 'EPSG:5186'
        
        self.apply_to_all = self.apply_all_checkbox.isChecked()
        self.accept()
    
    def skip_current(self) -> None:
        """현재 파일 건너뛰기"""
        self.skip_file = True
        self.accept()


class BatchProcessor(QThread):
    """배치 처리 워커 스레드"""
    
    # 시그널 정의
    progress_updated = pyqtSignal(int, str)  # 진행률, 메시지
    file_processed = pyqtSignal(dict)        # 파일 처리 결과
    crs_request = pyqtSignal(str, object)    # 좌표계 요청 (파일명, 콜백)
    batch_completed = pyqtSignal(str)        # 완료 (결과 파일 경로)
    error_occurred = pyqtSignal(str, str)    # 오류 (파일명, 오류메시지)
    
    def __init__(self, shp_folder: str, xlsx_folder: str, output_excel: str, model_path: str) -> None:
        super().__init__()
        self.shp_folder = Path(shp_folder)
        self.xlsx_folder = Path(xlsx_folder)
        self.output_excel = output_excel
        self.model_path = model_path
        
        self.shp_files: List[Path] = []
        self.results: List[Dict] = []
        self.is_running = True
        self.default_crs: Optional[str] = None  # 전체 적용할 기본 CRS
        
        # 좌표계 매핑 (지역명 기반)
        self.crs_mapping = {
            # 중부원점 (EPSG:5186)
            '서울': 'EPSG:5186', '인천': 'EPSG:5186', '대전': 'EPSG:5186',
            '광주': 'EPSG:5186', '세종': 'EPSG:5186', '경기': 'EPSG:5186',
            '충청': 'EPSG:5186', '전라': 'EPSG:5186', '전북': 'EPSG:5186',
            '제주': 'EPSG:5186',
            
            # 동부원점 (EPSG:5187)
            '부산': 'EPSG:5187', '대구': 'EPSG:5187', '울산': 'EPSG:5187',
            '경상': 'EPSG:5187', '경북': 'EPSG:5187', '경남': 'EPSG:5187',
        }
        
        # 강원도 동부 지역
        self.gangwon_east = {'속초', '고성', '양양', '강릉', '동해', '삼척', '태백'}
    
    def run(self) -> None:
        """배치 처리 실행"""
        try:
            # SHP 파일 수집
            self.collect_shp_files()
            
            if not self.shp_files:
                self.error_occurred.emit("", "SHP 파일을 찾을 수 없습니다.")
                return
            
            total_files = len(self.shp_files)
            self.progress_updated.emit(0, f"총 {total_files}개 파일 처리 시작...")
            
            # 각 파일 처리
            for i, shp_file in enumerate(self.shp_files):
                if not self.is_running:
                    break
                
                try:
                    # 진행률 업데이트
                    progress = int(i * 100 / total_files)
                    file_name = shp_file.name
                    self.progress_updated.emit(progress, f"[{i+1}/{total_files}] {file_name} 처리 중...")
                    
                    # 파일 처리
                    result = self.process_single_file(str(shp_file))
                    
                    if result:
                        self.results.append(result)
                        self.file_processed.emit(result)
                    
                    # 메모리 정리
                    gc.collect()
                    
                except Exception as e:
                    logger.error(f"파일 처리 오류 {shp_file}: {e}")
                    self.error_occurred.emit(shp_file.name, str(e))
                    continue
            
            # 결과 저장
            if self.results:
                self.save_results()
                self.progress_updated.emit(100, f"완료! 총 {len(self.results)}개 파일 처리됨")
                self.batch_completed.emit(self.output_excel)
            else:
                self.error_occurred.emit("", "처리된 파일이 없습니다.")
                
        except Exception as e:
            self.error_occurred.emit("", f"배치 처리 오류: {str(e)}")
    
    def collect_shp_files(self) -> None:
        """SHP 파일 수집"""
        self.shp_files = list(self.shp_folder.glob("*.shp"))
        logger.info(f"발견된 SHP 파일: {len(self.shp_files)}개")
    
    def detect_crs_auto(self, shp_file: str) -> Optional[str]:
        """자동 좌표계 감지"""
        try:
            # 1. SHP 파일에서 CRS 직접 읽기
            gdf = gpd.read_file(shp_file)
            if gdf.crs and gdf.crs.to_epsg() in [5186, 5187]:
                detected_crs = f"EPSG:{gdf.crs.to_epsg()}"
                logger.info(f"CRS 자동 감지: {Path(shp_file).name} → {detected_crs}")
                return detected_crs
            
            # 2. 파일명에서 지역명 추출하여 매핑
            file_name = Path(shp_file).name.lower()
            for region, crs in self.crs_mapping.items():
                if region in file_name:
                    logger.info(f"지역명 기반 CRS 매핑: {Path(shp_file).name} → {crs} ({region})")
                    return crs
            
            # 3. 강원도 특별 처리
            if '강원' in file_name:
                for east_city in self.gangwon_east:
                    if east_city in file_name:
                        return 'EPSG:5187'
                return 'EPSG:5186'  # 서부 강원도
            
        except Exception as e:
            logger.warning(f"CRS 자동 감지 실패: {Path(shp_file).name} - {e}")
        
        return None
    
    def process_single_file(self, shp_file: str) -> Optional[Dict]:
        """단일 파일 처리"""
        file_name = Path(shp_file).name
        
        try:
            # 1. 좌표계 결정
            crs = self.default_crs or self.detect_crs_auto(shp_file)
            
            if not crs:
                # 사용자에게 좌표계 요청
                crs = self.request_crs_from_user(shp_file)
                if not crs:  # 건너뛰기 또는 취소
                    logger.info(f"파일 건너뛰기: {file_name}")
                    return None
            
            # 2. AI 분석 수행
            ai_count = self.analyze_with_ai(shp_file, crs)
            
            # 3. XLSX 파일에서 실제 점 개수 추출
            xlsx_count = self.get_xlsx_point_count(shp_file)
            
            # 4. 결과 구성
            result = {
                'file_name': file_name,
                'ai_count': ai_count,
                'xlsx_count': xlsx_count,
                'crs': crs,
                'difference': ai_count - xlsx_count if xlsx_count is not None else None,
                'status': 'OK'
            }
            
            logger.info(f"처리 완료: {file_name} - AI:{ai_count}, 실제:{xlsx_count}, CRS:{crs}")
            return result
            
        except Exception as e:
            logger.error(f"파일 처리 실패: {file_name} - {e}")
            return {
                'file_name': file_name,
                'ai_count': 0,
                'xlsx_count': None,
                'crs': None,
                'difference': None,
                'status': f'ERROR: {str(e)[:50]}'
            }
    
    def request_crs_from_user(self, shp_file: str) -> Optional[str]:
        """사용자에게 좌표계 요청 (동기 처리)"""
        result_crs: Optional[str] = None
        dialog_finished = False
        
        def handle_dialog_result(dialog: CRSSelectionDialog) -> None:
            nonlocal result_crs, dialog_finished
            if dialog.skip_file:
                result_crs = None
            else:
                result_crs = dialog.selected_crs
                if dialog.apply_to_all:
                    self.default_crs = result_crs
            dialog_finished = True
        
        # 메인 스레드에서 다이얼로그 표시
        self.crs_request.emit(shp_file, handle_dialog_result)
        
        # 다이얼로그 완료까지 대기
        while not dialog_finished and self.is_running:
            self.msleep(100)
        
        return result_crs
    
    def detect_heuristic_intersections(self, skeleton):
        """간단한 휴리스틱 교차점 검출"""
        try:
            # 폴리곤 꼭짓점에서 각도 변화가 큰 지점을 교차점으로 판단
            intersections = []
            
            if len(skeleton) < 3:
                return intersections
            
            for i in range(1, len(skeleton) - 1):
                prev_point = skeleton[i-1]
                curr_point = skeleton[i]
                next_point = skeleton[i+1]
                
                # 벡터 계산
                v1 = np.array([curr_point[0] - prev_point[0], curr_point[1] - prev_point[1]])
                v2 = np.array([next_point[0] - curr_point[0], next_point[1] - curr_point[1]])
                
                # 벡터 길이 체크
                len_v1 = np.linalg.norm(v1)
                len_v2 = np.linalg.norm(v2)
                
                if len_v1 > 1e-6 and len_v2 > 1e-6:
                    # 정규화
                    v1_norm = v1 / len_v1
                    v2_norm = v2 / len_v2
                    
                    # 각도 계산
                    dot_product = np.dot(v1_norm, v2_norm)
                    dot_product = np.clip(dot_product, -1.0, 1.0)
                    angle = np.arccos(dot_product)
                    
                    # 급격한 방향 변화 (60도 이상)가 있으면 교차점으로 판단
                    if angle > np.pi / 3:  # 60도
                        intersections.append((float(curr_point[0]), float(curr_point[1])))
            
            return intersections[:5]  # 최대 5개로 제한
            
        except Exception as e:
            logger.warning(f"교차점 검출 오류: {e}")
            return []
    
    def analyze_with_ai(self, shp_file: str, crs: str) -> int:
        """🚀 Process3 완전 복제: 지구계로 도로망 클립 → 스켈레톤 추출 → AI 분석"""
        try:
            # ===== Process3의 정확한 방식: DistrictRoadClipper 사용 =====
            try:
                from src.core.district_road_clipper import DistrictRoadClipper
                from src.core.skeleton_extractor import SkeletonExtractor
            except ImportError:
                logger.error("핵심 모듈을 찾을 수 없음: DistrictRoadClipper 또는 SkeletonExtractor")
                return self.basic_analysis(shp_file)
            
            import tempfile
            
            district_clipper = DistrictRoadClipper()
            skeleton_extractor = SkeletonExtractor()
            
            # 1. 지구계 파일로 도로망 클립 (Process3와 동일)
            logger.info(f"🔍 지구계 파일로 도로망 클립 중: {Path(shp_file).name}")
            results = district_clipper.process_district_file(
                shp_file,
                target_crs=crs,
                auto_find_road=True
            )
            
            if not results['success']:
                logger.warning(f"❌ 도로망 클립 실패: {results.get('error', '알 수 없는 오류')}")
                logger.warning("기본 분석으로 대체")
                return self.basic_analysis(shp_file)
            
            total_final_points = 0
            
            # 2. 각 폴리곤별로 클립된 도로망 처리 (Process3와 동일)
            for idx, polygon_info in enumerate(results['polygons']):
                clipped_road = polygon_info.get('clipped_road')
                
                if clipped_road is None or clipped_road.empty:
                    logger.warning(f"폴리곤 {idx+1}: 클립된 도로망이 없음")
                    continue
                
                temp_path = None
                try:
                    # 3. 클립된 도로망을 임시 파일로 저장 (더 안전한 방식)
                    import tempfile
                    import os
                    
                    # 임시 디렉토리에 고유한 파일명 생성
                    temp_dir = tempfile.gettempdir()
                    temp_filename = f"clipped_road_{idx}_{os.getpid()}_{hash(str(clipped_road.geometry.iloc[0]))%10000}.shp"
                    temp_path = os.path.join(temp_dir, temp_filename)
                    
                    # 클립된 도로망 저장
                    clipped_road.to_file(temp_path)
                    
                    # 파일이 제대로 저장되었는지 확인
                    if not os.path.exists(temp_path):
                        logger.error(f"❌ 폴리곤 {idx+1}: 임시 파일 저장 실패: {temp_path}")
                        continue
                    
                    logger.info(f"📂 폴리곤 {idx+1}: 클립된 도로망 저장 완료 ({len(clipped_road)}개 세그먼트) → {temp_path}")
                    
                    # 4. SkeletonExtractor로 도로 스켈레톤 추출 (Process3의 핵심!)
                    logger.info(f"🦴 폴리곤 {idx+1}: 스켈레톤 추출 시작...")
                    skeleton, intersections = skeleton_extractor.process_shapefile(temp_path)
                    
                    if not skeleton or len(skeleton) < 3:
                        logger.warning(f"폴리곤 {idx+1}: 스켈레톤이 없음")
                        continue
                    
                    logger.info(f"🦴 폴리곤 {idx+1}: 스켈레톤 추출 완료 ({len(skeleton)}개 점, {len(intersections)}개 교차점)")
                    
                    # ===== Process3의 AI 분석 로직 적용 =====
                    
                    # 5-1. 휴리스틱 교차점은 스켈레톤에서 추출된 것 사용
                    canvas_points = {
                        'intersection': [(float(x), float(y)) for x, y in intersections],
                        'curve': [],
                        'endpoint': []
                    }
                    
                    # 5-2. 휴리스틱 끝점 검출 (스켈레톤 기반)
                    endpoints = self.detect_heuristic_endpoints(skeleton)
                    canvas_points['endpoint'] = endpoints
                    
                    # 5-3. 도로 경계선 기반 커브점 검출 (스켈레톤 기반)
                    curves = self.detect_boundary_based_curves(
                        skeleton,
                        sample_distance=15.0,
                        curvature_threshold=0.20,
                        road_buffer=3.0,
                        cluster_radius=20.0
                    )
                    
                    # 5-4. 교차점 근처 커브점 제거
                    filtered_curves = self.remove_curves_near_intersections(
                        curves, intersections, threshold=10.0
                    )
                    canvas_points['curve'] = filtered_curves
                    
                    logger.info(f"📍 폴리곤 {idx+1} 기본 분석: 교차점={len(intersections)}, 커브={len(filtered_curves)}, 끝점={len(endpoints)}")
                    
                    # ===== Process3의 핵심: 다단계 최적화 실행 =====
                    total_removed, optimized_final_points = self.run_multi_stage_optimization(canvas_points, skeleton)
                    
                    # ===== Process3와 동일한 최종 점 개수 =====
                    polygon_final_count = optimized_final_points
                    total_final_points += polygon_final_count
                    
                    logger.info(f"🎯 폴리곤 {idx+1} Process3 완료: 도로클립→스켈레톤→AI분석 → 최종 {polygon_final_count}개 점 (최적화: -{total_removed}개)")
                    
                except Exception as e:
                    logger.error(f"❌ 폴리곤 {idx+1} 처리 실패: {e}")
                    import traceback
                    logger.error(f"📋 오류 상세:\n{traceback.format_exc()}")
                    continue
                finally:
                    # 임시 파일 정리
                    if temp_path:
                        try:
                            Path(temp_path).unlink(missing_ok=True)
                            for ext in ['.shx', '.dbf', '.cpg', '.prj']:
                                Path(temp_path.replace('.shp', ext)).unlink(missing_ok=True)
                        except:
                            pass
            
            logger.info(f"✅ Process3 완전 복제 성공: {Path(shp_file).name} - 도로망 클립 기반 최종 {total_final_points}개 점")
            return total_final_points
            
        except Exception as e:
            logger.error(f"🚨 AI 분석 실패: {Path(shp_file).name}")
            logger.error(f"🔍 오류 타입: {type(e).__name__}")
            logger.error(f"🔍 오류 메시지: {str(e)}")
            import traceback
            logger.error(f"📋 전체 스택 트레이스:\n{traceback.format_exc()}")
            
            logger.info(f"🔄 기본 분석으로 폴백 시도: {Path(shp_file).name}")
            return self.basic_analysis(shp_file)
    
    def basic_analysis(self, shp_file: str) -> int:
        """기본 분석 방법 (PipelineManager 없을 때)"""
        logger.info(f"🔧 기본 분석 시작: {Path(shp_file).name}")
        
        try:
            # 간단한 기하학적 분석
            logger.info(f"📂 SHP 파일 읽기 시도: {Path(shp_file).name}")
            gdf = gpd.read_file(shp_file)
            logger.info(f"✅ SHP 파일 읽기 성공: {len(gdf)}개 레코드")
            
            if gdf.empty:
                logger.warning(f"⚠️ 빈 SHP 파일: {Path(shp_file).name}")
                return 0
            
            # 폴리곤의 꼭짓점 개수 기반 추정
            total_vertices = 0
            valid_geoms = 0
            
            for i, geom in enumerate(gdf.geometry):
                if geom and not geom.is_empty:
                    valid_geoms += 1
                    if hasattr(geom, 'exterior'):
                        coords_count = len(list(geom.exterior.coords))
                        total_vertices += coords_count
                        logger.debug(f"  폴리곤 {i+1}: {coords_count}개 좌표")
                    elif hasattr(geom, 'geoms'):
                        for j, sub_geom in enumerate(geom.geoms):
                            if hasattr(sub_geom, 'exterior'):
                                coords_count = len(list(sub_geom.exterior.coords))
                                total_vertices += coords_count
                                logger.debug(f"  멀티폴리곤 {i+1}-{j+1}: {coords_count}개 좌표")
                else:
                    logger.warning(f"  폴리곤 {i+1}: 유효하지 않은 geometry")
            
            # 추정 공식 (경험적)
            estimated_points = max(1, total_vertices // 20)  # 20개 꼭짓점당 1개 특징점 추정
            
            logger.info(f"🔢 기본 분석 결과: {Path(shp_file).name}")
            logger.info(f"  - 유효한 geometry: {valid_geoms}개")
            logger.info(f"  - 총 꼭짓점: {total_vertices}개")
            logger.info(f"  - 추정 점 개수: {estimated_points}개")
            
            return estimated_points
            
        except Exception as e:
            logger.error(f"🚨 기본 분석 실패: {Path(shp_file).name}")
            logger.error(f"🔍 오류 타입: {type(e).__name__}")
            logger.error(f"🔍 오류 메시지: {str(e)}")
            import traceback
            logger.error(f"📋 스택 트레이스:\n{traceback.format_exc()}")
            return 0  # 오류 시 0 반환
    
    def get_xlsx_point_count(self, shp_file: str) -> Optional[int]:
        """XLSX 파일에서 실제 점 개수 추출"""
        try:
            # SHP 파일명에서 '_현황_폴리곤' 제거하여 XLSX 파일명 생성
            shp_name = Path(shp_file).stem
            xlsx_name = shp_name.replace('_현황_폴리곤', '')
            xlsx_file = self.xlsx_folder / f"{xlsx_name}.xlsx"
            
            if not xlsx_file.exists():
                # 다른 확장자도 시도
                for ext in ['.XLSX', '.xls', '.XLS']:
                    xlsx_alt = self.xlsx_folder / f"{xlsx_name}{ext}"
                    if xlsx_alt.exists():
                        xlsx_file = xlsx_alt
                        break
                else:
                    logger.warning(f"XLSX 파일 없음: {xlsx_name} (원본: {shp_name})")
                    return None
            
            # XLSX 파일 읽기 및 중복 제거
            df = pd.read_excel(xlsx_file, header=None)
            
            # Y(1열), X(2열) 좌표 추출
            valid_points = []
            for _, row in df.iterrows():
                try:
                    y = float(row[0])  # 1열 (Y)
                    x = float(row[1])  # 2열 (X)
                    valid_points.append((x, y))
                except (ValueError, IndexError):
                    continue
            
            # 중복 좌표 제거 (set을 사용하여 중복 제거)
            unique_points = list(set(valid_points))
            original_count = len(valid_points)
            unique_count = len(unique_points)
            
            if original_count != unique_count:
                duplicates_removed = original_count - unique_count
                logger.info(f"📍 중복 제거: {xlsx_file.name} - 원본:{original_count}개 → 유효:{unique_count}개 (중복 제거:{duplicates_removed}개)")
            else:
                logger.info(f"📍 중복 없음: {xlsx_file.name} - {unique_count}개")
            
            return unique_count
            
        except Exception as e:
            logger.error(f"XLSX 읽기 오류: {shp_file} - {e}")
            return None
    
    def save_results(self) -> None:
        """결과를 엑셀 파일로 저장"""
        try:
            df = pd.DataFrame(self.results)
            
            # 컬럼 순서 및 이름 설정
            df = df[['file_name', 'ai_count', 'xlsx_count', 'difference', 'crs', 'status']]
            df.columns = ['SHP파일명', 'AI점개수', '실제점개수', '차이', '좌표계', '상태']
            
            # 엑셀 저장
            try:
                # pandas DataFrame을 엑셀로 저장
                df.to_excel(self.output_excel, sheet_name='분석결과', index=False, engine='openpyxl')
                
                # 워크시트 서식 설정을 위해 openpyxl 직접 사용
                from openpyxl import load_workbook
                wb = load_workbook(self.output_excel)
                ws = wb['분석결과']
                
                # 컬럼 너비 자동 조정
                from openpyxl.utils import get_column_letter
                for idx, column in enumerate(ws.columns, 1):
                    max_length = 0
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    
                    column_letter = get_column_letter(idx)
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(self.output_excel)
                wb.close()
            except ImportError:
                # openpyxl이 없으면 기본 저장만
                df.to_excel(self.output_excel, sheet_name='분석결과', index=False)
            
            logger.info(f"결과 저장 완료: {self.output_excel}")
            
        except Exception as e:
            logger.error(f"결과 저장 오류: {e}")
    
    def stop(self) -> None:
        """처리 중단"""
        self.is_running = False
    
    # ========== Process3와 동일한 분석 함수들 ==========
    
    def detect_heuristic_endpoints(self, skeleton, road_bounds=None):
        """휴리스틱 끝점 검출 - 지구계 경계 근처의 도로 끝 (Process3와 동일)"""
        if not skeleton:
            return []
        
        endpoints = []
        
        # 스켈레톤 포인트들의 경계 계산
        if skeleton:
            x_coords = [pt[0] for pt in skeleton if len(pt) >= 2]
            y_coords = [pt[1] for pt in skeleton if len(pt) >= 2]
            
            if x_coords and y_coords:
                min_x, max_x = min(x_coords), max(x_coords)
                min_y, max_y = min(y_coords), max(y_coords)
                
                # 경계로부터의 거리 임계값 (30m)
                threshold = 30.0
                
                for i, point in enumerate(skeleton):
                    if len(point) < 2:
                        continue
                    
                    x, y = float(point[0]), float(point[1])
                    
                    # 경계와의 거리 계산
                    dist_to_boundary = min(
                        x - min_x, max_x - x,  # 좌우 경계
                        y - min_y, max_y - y   # 상하 경계
                    )
                    
                    # 경계 근처이고 연결된 점이 적으면 끝점
                    if dist_to_boundary < threshold:
                        # 주변 연결점 개수 확인
                        connected_count = 0
                        for other_point in skeleton:
                            if len(other_point) < 2:
                                continue
                            other_x, other_y = float(other_point[0]), float(other_point[1])
                            if (x, y) != (other_x, other_y):
                                dist = np.sqrt((x - other_x)**2 + (y - other_y)**2)
                                if dist < 50:  # 50m 이내 연결점
                                    connected_count += 1
                        
                        # 연결점이 2개 이하면 끝점으로 판단
                        if connected_count <= 2:
                            endpoints.append((x, y))
        
        logger.info(f"🔚 휴리스틱 끝점 검출: {len(endpoints)}개")
        return endpoints
    
    def detect_boundary_based_curves(self, skeleton, sample_distance=15.0, curvature_threshold=0.20, 
                                   road_buffer=3.0, cluster_radius=20.0):
        """도로 경계선 기반 커브점 검출 (Process3와 동일)"""
        if not skeleton or len(skeleton) < 5:
            logger.info("스켈레톤이 너무 짧아 경계선 기반 커브 검출 불가")
            return []
        
        try:
            from shapely.geometry import LineString, Point
            
            # 전체 스켈레톤을 하나의 도로망으로 통합
            skeleton_coords = [(float(p[0]), float(p[1])) for p in skeleton if len(p) >= 2]
            if len(skeleton_coords) < 2:
                return []
            
            # 연속된 좌표들을 LineString으로 변환
            skeleton_line = LineString(skeleton_coords)
            
            # 통합된 도로에 버퍼 적용
            road_polygon = skeleton_line.buffer(road_buffer)
            
            # 복잡한 도로 형태 처리
            if road_polygon.geom_type == 'Polygon':
                boundaries = [road_polygon.exterior]
            elif road_polygon.geom_type == 'MultiPolygon':
                boundaries = []
                geoms = getattr(road_polygon, 'geoms', [])
                for poly in geoms:
                    if hasattr(poly, 'exterior'):
                        boundaries.append(poly.exterior)
            else:
                logger.warning(f"예상치 못한 geometry 타입: {road_polygon.geom_type}")
                return []
            
            # 모든 경계선에서 커브점 검출
            all_curvature_points = []
            
            for boundary in boundaries:
                total_length = boundary.length
                if total_length < sample_distance:
                    continue
            
                # 각 경계선을 따라 샘플링
                num_samples = max(10, int(total_length / sample_distance))
                
                for i in range(num_samples):
                    distance = (i * sample_distance) % total_length
                    
                    # 곡률 계산
                    curvature = self.calculate_curvature_at_distance(boundary, distance, sample_distance)
                    
                    if curvature > curvature_threshold:
                        point = boundary.interpolate(distance)
                        all_curvature_points.append({
                            'point': (point.x, point.y),
                            'curvature': curvature
                        })
            
            # 군집화
            if len(all_curvature_points) < 2:
                final_curves = [cp['point'] for cp in all_curvature_points]
            else:
                try:
                    from sklearn.cluster import DBSCAN
                    points = np.array([cp['point'] for cp in all_curvature_points])
                    clustering = DBSCAN(eps=cluster_radius, min_samples=2)
                    labels = clustering.fit_predict(points)
                    
                    final_curves = []
                    for cluster_id in set(labels):
                        if cluster_id == -1:  # 노이즈 포인트들
                            cluster_points = points[labels == cluster_id]
                            final_curves.extend([(p[0], p[1]) for p in cluster_points])
                        else:  # 클러스터들
                            cluster_points = points[labels == cluster_id]
                            cluster_center = cluster_points.mean(axis=0)
                            final_curves.append((cluster_center[0], cluster_center[1]))
                except ImportError:
                    # sklearn 없으면 원본 포인트 사용
                    final_curves = [cp['point'] for cp in all_curvature_points]
            
            # 스켈레톤에 가장 가까운 점으로 조정
            skeleton_adjusted_curves = []
            for curve_point in final_curves:
                closest_skeleton_point = self.find_closest_skeleton_point(curve_point, skeleton)
                if closest_skeleton_point:
                    skeleton_adjusted_curves.append(closest_skeleton_point)
            
            logger.info(f"🔄 경계선 기반 커브점 검출: {len(skeleton_adjusted_curves)}개")
            return skeleton_adjusted_curves
            
        except Exception as e:
            logger.error(f"경계선 커브 검출 오류: {e}")
            return []
    
    def calculate_curvature_at_distance(self, boundary, distance, window=20.0):
        """특정 거리에서의 곡률 계산 (Process3와 동일)"""
        try:
            total_length = boundary.length
            
            # 윈도우 크기 조정
            half_window = window / 2
            start_distance = max(0, distance - half_window)
            end_distance = min(total_length, distance + half_window)
            
            # 세 점 샘플링
            p1 = boundary.interpolate(start_distance)
            p2 = boundary.interpolate(distance)
            p3 = boundary.interpolate(end_distance)
            
            # 곡률 계산 (삼각형 외접원의 곡률)
            x1, y1 = p1.x, p1.y
            x2, y2 = p2.x, p2.y
            x3, y3 = p3.x, p3.y
            
            # 벡터 계산
            v1 = np.array([x2 - x1, y2 - y1])
            v2 = np.array([x3 - x2, y3 - y2])
            
            # 길이 체크
            len_v1 = np.linalg.norm(v1)
            len_v2 = np.linalg.norm(v2)
            
            if len_v1 < 1e-6 or len_v2 < 1e-6:
                return 0.0
            
            # 정규화
            v1_norm = v1 / len_v1
            v2_norm = v2 / len_v2
            
            # 각도 변화 계산
            cross_product = np.cross(v1_norm, v2_norm)
            dot_product = np.dot(v1_norm, v2_norm)
            
            # 각도 변화를 곡률로 변환
            angle_change = np.arctan2(abs(cross_product), dot_product)
            avg_length = (len_v1 + len_v2) / 2
            
            curvature = angle_change / max(avg_length, 1e-6)
            
            return curvature
            
        except Exception:
            return 0.0
    
    def find_closest_skeleton_point(self, curve_point, skeleton):
        """커브점에 가장 가까운 스켈레톤 점 찾기 (Process3와 동일)"""
        if not skeleton:
            return None
        
        min_dist = float('inf')
        closest_point = None
        
        for skel_point in skeleton:
            if len(skel_point) < 2:
                continue
            
            dist = np.sqrt((curve_point[0] - skel_point[0])**2 + 
                          (curve_point[1] - skel_point[1])**2)
            
            if dist < min_dist:
                min_dist = dist
                closest_point = (float(skel_point[0]), float(skel_point[1]))
        
        return closest_point
    
    def remove_curves_near_intersections(self, curves, intersections, threshold=10.0):
        """교차점 근처 커브점 제거 (Process3와 동일)"""
        if not curves or not intersections:
            return curves
        
        filtered_curves = []
        
        for curve in curves:
            near_intersection = False
            
            for intersection in intersections:
                if len(intersection) < 2:
                    continue
                
                dist = np.sqrt((curve[0] - intersection[0])**2 + 
                              (curve[1] - intersection[1])**2)
                
                if dist <= threshold:
                    near_intersection = True
                    break
            
            if not near_intersection:
                filtered_curves.append(curve)
        
        logger.info(f"교차점 근처 커브점 제거: {len(curves)} → {len(filtered_curves)}개")
        return filtered_curves

    def run_multi_stage_optimization(self, canvas_points, skeleton):
        """Process3와 동일한 다단계 점 최적화 실행"""
        try:
            # 1단계: 지능형 클러스터링 (20m 반경 중요도 기반)
            stage1_removed = self.intelligent_clustering_optimization(canvas_points, skeleton)
            
            # 2단계: 가까운 점 클러스터링 삭제 (15m)
            stage2_removed = self.remove_clustered_points(canvas_points, 15.0)
            
            # 3단계: 연결성 기반 커브점 1개 삭제
            stage3_removed = 1 if self.remove_one_curve_point_by_connectivity(canvas_points) else 0
            
            # 4단계: 자동 끝점 정리 (경계 근처 고립점)
            stage4_removed = self.auto_remove_road_endpoints(canvas_points, skeleton)
            
            # 5단계: 중복점 필터링
            self.filter_overlapping_points(canvas_points, skeleton)
            
            total_removed = stage1_removed + stage2_removed + stage3_removed + stage4_removed
            final_points = sum(len(canvas_points.get(cat, [])) for cat in ['intersection', 'curve', 'endpoint'])
            
            return total_removed, final_points
            
        except Exception as e:
            logger.warning(f"다단계 최적화 실패: {e}")
            return 0, sum(len(canvas_points.get(cat, [])) for cat in ['intersection', 'curve', 'endpoint'])
    
    def intelligent_clustering_optimization(self, canvas_points, skeleton):
        """20m 반경 지능형 클러스터링 - 중요도 기반 점 선택"""
        try:
            from sklearn.cluster import DBSCAN
            
            all_points = []
            for category in ['intersection', 'curve', 'endpoint']:
                for point in canvas_points.get(category, []):
                    all_points.append({
                        'coord': point,
                        'category': category
                    })
            
            if len(all_points) < 2:
                return 0
            
            # 중요도 계산
            skeleton_array = np.array(skeleton) if skeleton else np.array([])
            for point_data in all_points:
                point_data['importance'] = self.calculate_point_importance(
                    point_data['coord'], point_data['category'], skeleton_array)
            
            # DBSCAN 클러스터링
            coords = np.array([p['coord'] for p in all_points])
            clustering = DBSCAN(eps=20.0, min_samples=1).fit(coords)
            
            optimized_points = {'intersection': [], 'curve': [], 'endpoint': []}
            clusters = {}
            
            for i, label in enumerate(clustering.labels_):
                if label not in clusters:
                    clusters[label] = []
                clusters[label].append(all_points[i])
            
            removed_count = 0
            
            for cluster_points in clusters.values():
                if len(cluster_points) == 1:
                    point = cluster_points[0]
                    optimized_points[point['category']].append(point['coord'])
                else:
                    # 중요도 기반 최적 점 선택
                    best_point = max(cluster_points, key=lambda p: p['importance'])
                    optimized_points[best_point['category']].append(best_point['coord'])
                    removed_count += len(cluster_points) - 1
            
            # 결과 적용
            for category in ['intersection', 'curve', 'endpoint']:
                canvas_points[category] = optimized_points[category]
            
            return removed_count
            
        except ImportError:
            logger.warning("scikit-learn 없음 - 기본 클러스터링 사용")
            return 0
        except Exception as e:
            logger.warning(f"지능형 클러스터링 실패: {e}")
            return 0
    
    def calculate_point_importance(self, point_coord, category, skeleton_array):
        """점의 중요도 계산 - 스켈레톤 밀도와 카테고리 기반"""
        x, y = point_coord
        importance_score = 0.0
        
        # 1. 카테고리 기반 기본 중요도
        category_weights = {
            'intersection': 10.0,  # 교차점이 가장 중요
            'endpoint': 7.0,       # 끝점이 두 번째 중요
            'curve': 5.0           # 커브점이 세 번째 중요
        }
        importance_score += category_weights.get(category, 0.0)
        
        # 2. 스켈레톤 밀도 기반 중요도
        if len(skeleton_array) > 0:
            distances = np.sqrt(np.sum((skeleton_array - np.array([x, y]))**2, axis=1))
            
            # 다양한 반경의 근처 점 개수
            nearby_count_50m = np.sum(distances <= 50.0)
            nearby_count_30m = np.sum(distances <= 30.0)
            nearby_count_10m = np.sum(distances <= 10.0)
            
            # 밀도 점수 계산
            density_score = (nearby_count_10m * 3.0 + 
                           nearby_count_30m * 2.0 + 
                           nearby_count_50m * 1.0)
            importance_score += density_score
        
        return importance_score
    
    def remove_clustered_points(self, canvas_points, distance_threshold=15.0):
        """가까운 점들 클러스터링 제거"""
        all_points = []
        for category in ['intersection', 'curve', 'endpoint']:
            for point in canvas_points.get(category, []):
                all_points.append({
                    'point': point,
                    'category': category
                })
        
        if len(all_points) < 2:
            return 0
        
        points_to_remove = []
        used_indices = set()
        
        for i, p1 in enumerate(all_points):
            if i in used_indices:
                continue
            
            nearby_points = []
            for j, p2 in enumerate(all_points[i+1:], i+1):
                if i != j and j not in used_indices:
                    try:
                        dist = np.sqrt(
                            (float(p1['point'][0]) - float(p2['point'][0]))**2 + 
                            (float(p1['point'][1]) - float(p2['point'][1]))**2
                        )
                        if dist <= distance_threshold:
                            nearby_points.append((j, p2, dist))
                    except:
                        continue
            
            if nearby_points:
                nearby_points.sort(key=lambda x: x[2])
                to_remove_idx, to_remove_point, _ = nearby_points[0]
                points_to_remove.append(to_remove_point)
                used_indices.add(to_remove_idx)
                used_indices.add(i)
        
        deleted_count = 0
        for point_info in points_to_remove:
            category = point_info['category']
            point = point_info['point']
            
            if point in canvas_points[category]:
                canvas_points[category].remove(point)
                deleted_count += 1
        
        return deleted_count
    
    def remove_one_curve_point_by_connectivity(self, canvas_points):
        """연결성 검사로 직선상 커브점 1개 삭제"""
        try:
            all_points = []
            for category in ['intersection', 'curve', 'endpoint']:
                for point in canvas_points.get(category, []):
                    all_points.append({
                        'point': point,
                        'category': category,
                        'coords': (float(point[0]), float(point[1]))
                    })
            
            if len(all_points) < 3:
                return False
            
            deletable_curves = []
            
            for i, p1 in enumerate(all_points):
                for j, p2 in enumerate(all_points):
                    if i >= j:
                        continue
                    
                    try:
                        from shapely.geometry import LineString, Point
                        line = LineString([p1['coords'], p2['coords']])
                        
                        if line.length < 10:
                            continue
                        
                        for k, p3 in enumerate(all_points):
                            if k == i or k == j or p3['category'] != 'curve':
                                continue
                            
                            point_dist_to_line = line.distance(Point(p3['coords']))
                            
                            if point_dist_to_line < 5.0:
                                segment_length = np.sqrt(
                                    (p1['coords'][0] - p2['coords'][0])**2 + 
                                    (p1['coords'][1] - p2['coords'][1])**2
                                )
                                
                                deletable_curves.append({
                                    'point_info': p3,
                                    'distance': segment_length,
                                })
                    except Exception:
                        continue
            
            deletable_curves.sort(key=lambda x: x['distance'])
            
            if deletable_curves:
                to_delete = deletable_curves[0]
                point_info = to_delete['point_info']
                
                if point_info['point'] in canvas_points['curve']:
                    canvas_points['curve'].remove(point_info['point'])
                    return True
            
            return False
            
        except Exception:
            return False
    
    def auto_remove_road_endpoints(self, canvas_points, skeleton):
        """자동 끝점 정리 - 경계 근처 고립된 끝점 제거"""
        if not skeleton:
            return 0
        
        removed_count = 0
        endpoints = canvas_points.get('endpoint', [])
        
        if not endpoints:
            return 0
        
        # 스켈레톤 경계 계산
        try:
            x_coords = [pt[0] for pt in skeleton if len(pt) >= 2]
            y_coords = [pt[1] for pt in skeleton if len(pt) >= 2]
            
            if not x_coords or not y_coords:
                return 0
            
            min_x, max_x = min(x_coords), max(x_coords)
            min_y, max_y = min(y_coords), max(y_coords)
            
            # 경계로부터 거리 임계값 (30m)
            threshold = 30.0
            
            endpoints_to_remove = []
            
            for endpoint in endpoints:
                try:
                    x, y = float(endpoint[0]), float(endpoint[1])
                    
                    # 경계와의 거리 계산
                    dist_to_boundary = min(
                        x - min_x, max_x - x,  # 좌우 경계
                        y - min_y, max_y - y   # 상하 경계
                    )
                    
                    # 경계 근처이고 연결된 점이 적으면 끝점 제거 후보
                    if dist_to_boundary < threshold:
                        # 주변 연결점 개수 확인
                        connected_count = 0
                        for other_point in skeleton:
                            if len(other_point) < 2:
                                continue
                            other_x, other_y = float(other_point[0]), float(other_point[1])
                            if (x, y) != (other_x, other_y):
                                dist = np.sqrt((x - other_x)**2 + (y - other_y)**2)
                                if dist < 50:  # 50m 이내 연결점
                                    connected_count += 1
                        
                        # 연결점이 2개 이하면 제거
                        if connected_count <= 2:
                            endpoints_to_remove.append(endpoint)
                except:
                    continue
            
            # 제거 실행
            for endpoint in endpoints_to_remove:
                if endpoint in canvas_points['endpoint']:
                    canvas_points['endpoint'].remove(endpoint)
                    removed_count += 1
            
            return removed_count
            
        except Exception:
            return 0
    
    def filter_overlapping_points(self, canvas_points, skeleton):
        """중복점 필터링 (하이브리드 필터)"""
        try:
            all_points = []
            point_roles = {}
            for category in ['intersection', 'curve', 'endpoint']:
                for point in canvas_points.get(category, []):
                    all_points.append(point)
                    point_roles[point] = category
            
            if len(all_points) < 2:
                return
            
            if not skeleton:
                return
            
            # 기본 중복 제거 (매우 가까운 점들)
            threshold_very_close = 5.0
            to_remove = []
            
            for i, p1 in enumerate(all_points):
                for j, p2 in enumerate(all_points[i+1:], i+1):
                    try:
                        dist = np.sqrt((p1[0] - p2[0])**2 + (p1[1] - p2[1])**2)
                        if dist < threshold_very_close:
                            # 중요도가 낮은 점 제거
                            if point_roles[p1] == 'curve' and point_roles[p2] != 'curve':
                                to_remove.append(p1)
                            elif point_roles[p2] == 'curve' and point_roles[p1] != 'curve':
                                to_remove.append(p2)
                            elif point_roles[p1] == 'endpoint' and point_roles[p2] == 'intersection':
                                to_remove.append(p1)
                            elif point_roles[p2] == 'endpoint' and point_roles[p1] == 'intersection':
                                to_remove.append(p2)
                    except:
                        continue
            
            # 제거 실행
            for point in set(to_remove):
                category = point_roles[point]
                if point in canvas_points[category]:
                    canvas_points[category].remove(point)
            
        except Exception:
            pass


class BatchSHPAnalyzer(QMainWindow):
    """메인 GUI 클래스"""
    
    def __init__(self) -> None:
        super().__init__()
        self.shp_folder = ""
        self.xlsx_folder = ""
        self.output_excel = ""
        self.model_path = ""
        
        self.processor: Optional[BatchProcessor] = None
        self.results: List[Dict] = []
        
        self.init_ui()
        self.check_models()
    
    def init_ui(self) -> None:
        """UI 초기화"""
        self.setWindowTitle("🤖 SHP 파일 배치 AI 분석기")
        self.setGeometry(100, 100, 900, 700)
        
        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout()
        central_widget.setLayout(layout)
        
        # 1. 모델 선택
        model_group = QGroupBox("1. AI 모델 선택")
        model_layout = QHBoxLayout()
        
        self.model_combo = QComboBox()
        model_layout.addWidget(QLabel("모델:"))
        model_layout.addWidget(self.model_combo)
        
        refresh_btn = QPushButton("새로고침")
        refresh_btn.clicked.connect(self.check_models)
        model_layout.addWidget(refresh_btn)
        
        model_group.setLayout(model_layout)
        layout.addWidget(model_group)
        
        # 2. 폴더 선택
        folder_group = QGroupBox("2. 폴더 선택")
        folder_layout = QGridLayout()
        
        # SHP 폴더
        folder_layout.addWidget(QLabel("SHP 폴더:"), 0, 0)
        self.shp_folder_label = QLabel("선택되지 않음")
        self.shp_folder_label.setStyleSheet("QLabel {background-color: #f8f9fa; padding: 5px; border: 1px solid #dee2e6;}")
        folder_layout.addWidget(self.shp_folder_label, 0, 1)
        
        shp_btn = QPushButton("📁 선택")
        shp_btn.clicked.connect(self.select_shp_folder)
        folder_layout.addWidget(shp_btn, 0, 2)
        
        # XLSX 폴더
        folder_layout.addWidget(QLabel("XLSX 폴더:"), 1, 0)
        self.xlsx_folder_label = QLabel("선택되지 않음")
        self.xlsx_folder_label.setStyleSheet("QLabel {background-color: #f8f9fa; padding: 5px; border: 1px solid #dee2e6;}")
        folder_layout.addWidget(self.xlsx_folder_label, 1, 1)
        
        xlsx_btn = QPushButton("📁 선택")
        xlsx_btn.clicked.connect(self.select_xlsx_folder)
        folder_layout.addWidget(xlsx_btn, 1, 2)
        
        # 출력 파일
        folder_layout.addWidget(QLabel("출력 엑셀:"), 2, 0)
        self.output_label = QLabel("results.xlsx")
        self.output_label.setStyleSheet("QLabel {background-color: #f8f9fa; padding: 5px; border: 1px solid #dee2e6;}")
        folder_layout.addWidget(self.output_label, 2, 1)
        
        output_btn = QPushButton("📄 설정")
        output_btn.clicked.connect(self.select_output_file)
        folder_layout.addWidget(output_btn, 2, 2)
        
        folder_group.setLayout(folder_layout)
        layout.addWidget(folder_group)
        
        # 3. 실행 버튼
        control_layout = QHBoxLayout()
        
        self.start_btn = QPushButton("🚀 분석 시작")
        self.start_btn.setStyleSheet("QPushButton {background-color: #28a745; color: white; font-weight: bold; padding: 10px;}")
        self.start_btn.clicked.connect(self.start_analysis)
        control_layout.addWidget(self.start_btn)
        
        self.stop_btn = QPushButton("⏹️ 중단")
        self.stop_btn.setStyleSheet("QPushButton {background-color: #dc3545; color: white; padding: 10px;}")
        self.stop_btn.clicked.connect(self.stop_analysis)
        self.stop_btn.setEnabled(False)
        control_layout.addWidget(self.stop_btn)
        
        layout.addLayout(control_layout)
        
        # 4. 진행 상황
        progress_group = QGroupBox("3. 진행 상황")
        progress_layout = QVBoxLayout()
        
        self.progress_bar = QProgressBar()
        progress_layout.addWidget(self.progress_bar)
        
        self.progress_label = QLabel("대기 중...")
        progress_layout.addWidget(self.progress_label)
        
        progress_group.setLayout(progress_layout)
        layout.addWidget(progress_group)
        
        # 5. 결과 테이블
        result_group = QGroupBox("4. 분석 결과")
        result_layout = QVBoxLayout()
        
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(6)
        self.result_table.setHorizontalHeaderLabels(['SHP파일명', 'AI점개수', '실제점개수', '차이', '좌표계', '상태'])
        
        # 테이블 설정
        header = self.result_table.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        
        result_layout.addWidget(self.result_table)
        
        # 통계 정보
        self.stats_label = QLabel("처리된 파일: 0개")
        result_layout.addWidget(self.stats_label)
        
        result_group.setLayout(result_layout)
        layout.addWidget(result_group)
        
        # 상태바
        self.statusBar().showMessage("준비")
    
    def check_models(self) -> None:
        """AI 모델 확인"""
        self.model_combo.clear()
        
        models_dir = Path("models")
        if models_dir.exists():
            model_files = list(models_dir.glob("*.pth"))
            if model_files:
                for model_file in model_files:
                    self.model_combo.addItem(model_file.name, str(model_file))
                self.statusBar().showMessage(f"{len(model_files)}개 모델 발견")
            else:
                self.model_combo.addItem("모델 없음", "")
                self.statusBar().showMessage("AI 모델이 없습니다.")
        else:
            self.model_combo.addItem("models 폴더 없음", "")
            self.statusBar().showMessage("models 폴더가 없습니다.")
    
    def select_shp_folder(self) -> None:
        """SHP 폴더 선택"""
        folder = QFileDialog.getExistingDirectory(self, "SHP 파일 폴더 선택")
        if folder:
            self.shp_folder = folder
            self.shp_folder_label.setText(folder)
            
            # SHP 파일 개수 확인
            shp_count = len(list(Path(folder).glob("*.shp")))
            self.statusBar().showMessage(f"SHP 파일 {shp_count}개 발견")
    
    def select_xlsx_folder(self) -> None:
        """XLSX 폴더 선택"""
        folder = QFileDialog.getExistingDirectory(self, "XLSX 파일 폴더 선택")
        if folder:
            self.xlsx_folder = folder
            self.xlsx_folder_label.setText(folder)
            
            # XLSX 파일 개수 확인
            xlsx_count = len(list(Path(folder).glob("*.xlsx"))) + len(list(Path(folder).glob("*.XLSX")))
            self.statusBar().showMessage(f"XLSX 파일 {xlsx_count}개 발견")
    
    def select_output_file(self) -> None:
        """출력 파일 선택"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "결과 엑셀 파일 저장", 
            "batch_analysis_results.xlsx",
            "Excel Files (*.xlsx)"
        )
        if file_path:
            self.output_excel = file_path
            self.output_label.setText(file_path)
    
    def start_analysis(self) -> None:
        """분석 시작"""
        # 입력 검증
        if not self.shp_folder:
            QMessageBox.warning(self, "경고", "SHP 폴더를 선택하세요.")
            return
        
        if not self.xlsx_folder:
            QMessageBox.warning(self, "경고", "XLSX 폴더를 선택하세요.")
            return
        
        if not self.output_excel:
            self.output_excel = str(Path.cwd() / "batch_analysis_results.xlsx")
            self.output_label.setText(self.output_excel)
        
        self.model_path = self.model_combo.currentData()
        if not self.model_path or not Path(self.model_path).exists():
            QMessageBox.warning(self, "경고", "유효한 AI 모델을 선택하세요.")
            return
        
        # UI 상태 변경
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.result_table.setRowCount(0)
        self.results.clear()
        
        # 처리 스레드 시작
        self.processor = BatchProcessor(
            self.shp_folder, 
            self.xlsx_folder, 
            self.output_excel,
            self.model_path
        )
        
        # 시그널 연결
        self.processor.progress_updated.connect(self.update_progress)
        self.processor.file_processed.connect(self.add_result)
        self.processor.crs_request.connect(self.handle_crs_request)
        self.processor.batch_completed.connect(self.on_batch_completed)
        self.processor.error_occurred.connect(self.on_error)
        
        self.processor.start()
        self.statusBar().showMessage("배치 분석 시작...")
    
    def stop_analysis(self) -> None:
        """분석 중단"""
        if self.processor:
            self.processor.stop()
            self.processor.wait()
        
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.statusBar().showMessage("분석 중단됨")
    
    def update_progress(self, value: int, message: str) -> None:
        """진행률 업데이트"""
        self.progress_bar.setValue(value)
        self.progress_label.setText(message)
    
    def add_result(self, result: Dict) -> None:
        """결과 추가"""
        self.results.append(result)
        
        # 테이블에 행 추가
        row = self.result_table.rowCount()
        self.result_table.insertRow(row)
        
        # 데이터 입력
        self.result_table.setItem(row, 0, QTableWidgetItem(result['file_name']))
        self.result_table.setItem(row, 1, QTableWidgetItem(str(result['ai_count'])))
        
        xlsx_count = result['xlsx_count']
        self.result_table.setItem(row, 2, QTableWidgetItem(str(xlsx_count) if xlsx_count is not None else "N/A"))
        
        difference = result['difference']
        diff_item = QTableWidgetItem(str(difference) if difference is not None else "N/A")
        if difference is not None:
            if difference > 0:
                diff_item.setBackground(QColor(255, 200, 200))  # 빨간색 (AI가 더 많음)
            elif difference < 0:
                diff_item.setBackground(QColor(200, 200, 255))  # 파란색 (실제가 더 많음)
            else:
                diff_item.setBackground(QColor(200, 255, 200))  # 초록색 (같음)
        
        self.result_table.setItem(row, 3, diff_item)
        self.result_table.setItem(row, 4, QTableWidgetItem(result['crs'] or "N/A"))
        self.result_table.setItem(row, 5, QTableWidgetItem(result['status']))
        
        # 통계 업데이트
        self.stats_label.setText(f"처리된 파일: {len(self.results)}개")
        
        # 스크롤 하단으로
        self.result_table.scrollToBottom()
    
    def handle_crs_request(self, shp_file: str, callback) -> None:
        """좌표계 요청 처리"""
        dialog = CRSSelectionDialog(shp_file, self)
        dialog.exec_()
        callback(dialog)
    
    def on_batch_completed(self, output_file: str) -> None:
        """배치 완료"""
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_bar.setValue(100)
        
        successful_count = len([r for r in self.results if r['status'] == 'OK'])
        total_count = len(self.results)
        
        QMessageBox.information(
            self, "배치 분석 완료",
            f"총 {total_count}개 파일 중 {successful_count}개 성공적으로 처리됨\n\n"
            f"결과 파일: {output_file}"
        )
        
        self.statusBar().showMessage(f"완료 - {successful_count}/{total_count} 파일 처리됨")
        
        # 결과 파일 열기 제안
        reply = QMessageBox.question(
            self, "결과 파일 열기", 
            "결과 엑셀 파일을 열어보시겠습니까?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.open_result_file(output_file)
    
    def open_result_file(self, file_path: str) -> None:
        """결과 파일 열기"""
        try:
            if sys.platform.startswith('win'):
                os.startfile(file_path)  # Windows
            elif sys.platform.startswith('darwin'):
                subprocess.run(['open', file_path], check=True)  # macOS
            else:
                subprocess.run(['xdg-open', file_path], check=True)  # Linux
        except (OSError, subprocess.CalledProcessError) as e:
            logger.error(f"파일 열기 실패: {e}")
            QMessageBox.warning(self, "오류", f"파일을 열 수 없습니다: {str(e)}")
    
    def on_error(self, file_name: str, error_msg: str) -> None:
        """오류 처리"""
        logger.error(f"오류 발생: {file_name} - {error_msg}")
        self.statusBar().showMessage(f"오류: {file_name} - {error_msg[:50]}")
    
    def closeEvent(self, event) -> None:
        """종료 시 처리"""
        if self.processor and self.processor.isRunning():
            reply = QMessageBox.question(
                self, "종료 확인",
                "배치 처리가 진행 중입니다. 종료하시겠습니까?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                self.stop_analysis()
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()


def main() -> None:
    """메인 실행 함수"""
    app = QApplication(sys.argv)
    app.setApplicationName("SHP 배치 AI 분석기")
    app.setOrganizationName("AI Road Analysis")
    
    # 고해상도 디스플레이 지원 (타입 체커 오류 방지를 위해 getattr 사용)
    try:
        aa_enable_high_dpi = getattr(Qt, 'AA_EnableHighDpiScaling', None)
        aa_use_high_dpi_pixmaps = getattr(Qt, 'AA_UseHighDpiPixmaps', None)
        if aa_enable_high_dpi is not None:
            app.setAttribute(aa_enable_high_dpi, True)
        if aa_use_high_dpi_pixmaps is not None:
            app.setAttribute(aa_use_high_dpi_pixmaps, True)
    except (AttributeError, TypeError):
        # Qt 버전이 낮거나 지원하지 않는 경우 무시
        pass
    
    # 메인 윈도우 생성
    window = BatchSHPAnalyzer()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == '__main__':
    main() 